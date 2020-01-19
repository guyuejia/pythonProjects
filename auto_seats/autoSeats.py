# -*- coding: utf-8 -*-
'''
@File    :   autoSeats.py    
@Contact :   hushishuai.fly@hotmail.com
@License :   (C)Copyright 2017-2018, Liugroup-NLPR-CASIA
@Modify Time      @Author    @Version    @Desciption
------------      -------    --------    -----------
2020/1/18 19:06   gujiayue      1.0         None
'''
import openpyxl

line1 = ["B2","D2","A2","E2"]
line2 = ["B4","D4","A4","E4"]
lines = ["B6","D6","A6","E6",
         "B8", "D8", "A8", "E8",
         "B10", "D10", "A10", "E10",
         "B12", "D12", "A12", "E12"]
#给合并单元格赋值，由于无法直接赋值，需要先拆分，再合并
def setMergeCell(ws,cellID,value):
    column = cellID[0]
    row = cellID[1:]
    nextRow = int(row) + 1
    #获取跟参数单元格合并的单元格ID
    nextCellID = column+str(nextRow)
    cellIDs = cellID + ":" + nextCellID
    #拆分单元格
    ws.unmerge_cells(cellIDs)
    #赋值
    ws[cellID] = value
    #合并
    ws.merge_cells(cellIDs)



#获取通讯录
wb = openpyxl.load_workbook("address_list.xlsx")
ws = wb.active
leader = []
vice_leaders = []
directors = []
staffs = []

rowNum = ws.max_row
columnNum = ws.max_column

for i in range(2,rowNum-1):
    if ws.cell(i, column=1).value == "处长" and  ws.cell(i, column=3).value == "是":
        leader.append(ws.cell(i, column=2).value)
        continue
    if ws.cell(i, column=1).value == "副处长" and  ws.cell(i, column=3).value == "是":
        vice_leaders.append(ws.cell(i, column=2).value)
        continue
    if ws.cell(i, column=1).value == "主管" and ws.cell(i, column=3).value == "是":
        directors.append(ws.cell(i, column=2).value)
        continue
    if ws.cell(i, column=1).value == "员工" and  ws.cell(i, column=3).value == "是":
        staffs.append(ws.cell(i, column=2).value)
        continue
wb.close()
print(leader)
print(vice_leaders)
print(directors)
print(staffs)

#打开会议室座位图
meeting_wb = openpyxl.load_workbook("meeting_room.xlsx")
meeting_ws = meeting_wb["Sheet1"]

is_leader ,is_vice_leaders, is_directors, is_staffs = False ,False,False,False
if len(leader) !=0:
    is_leader = True
if len(vice_leaders) !=0:
    is_vice_leaders = True
if len(directors) !=0:
    is_directors = True
if len(staffs) !=0:
    is_staffs = True
#给处长分配位置
if is_leader:
    setMergeCell(meeting_ws,line1[0],leader[0])
    #位置分配后，就删除
    line1.pop(0)
else:
    print("处长不参会")

#给副处长分配位置
if is_vice_leaders:
    for i in range(0,len(vice_leaders)):
        #永远都分配line1中的第一个位置，分配后就删除
        print(vice_leaders[i])
        setMergeCell(meeting_ws,line1[0],vice_leaders[i])
        line1.pop(0)
else:
    print("副处长不参会")

if is_directors:
    #首先判断处长或者副处长是否有，存在的话说明第一行被占用了
    if is_leader or is_vice_leaders:
        for i in range(0,len(directors)):
            setMergeCell(meeting_ws,line2[0],directors[i])
            line2.pop(0)
    #不存在的话，就直接分配第一行
    else:
        for i in range(0, len(directors)):
            setMergeCell(meeting_ws, line1[0], directors[i])
            line1.pop(0)
else:
    print("主管不参会")

if is_staffs:
    #先判断第一排是否有人,如果有再判断第二排是否会有人
    if is_leader or is_vice_leaders:
        print("第一排是有人做的")
        if not is_directors:
            print("第二排是无人做的")
            #如果不存在，则把第二行配置拼接到lines
            line2.extend(lines)
            lines = line2
        else:
            print("第二排是有人做的")
    #如果第一行就是空的，需要把所有位置拼接到一起
    else:
        print("第一排是无人做的")
        line1.extend(line2)
        line1.extend(lines)
        lines = line1
    #员工只需要按照所有的位置顺序坐即可
    for i in range(0, len(staffs)):
        print("分配位置：" + lines[0])
        setMergeCell(meeting_ws, lines[0], staffs[i])
        #meeting_ws[lines[0]] = staffs[i]
        lines.pop(0)

meeting_wb.save("meeting_room.xlsx")
meeting_wb.close()
